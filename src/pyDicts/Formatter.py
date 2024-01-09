from typing import Optional, Union, Tuple
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet import worksheet


class Formatter:
    def __init__(self, sht: worksheet, font: Optional[Font] = None, alignment: Optional[Alignment] = None,
                 fill: Optional[PatternFill] = None):
        self.sht = sht
        if font is None:
            font = Font(name='Arial Nova Cond', size=11, color="FFFFFF")

        if alignment is None:
            alignment = Alignment(horizontal='center', vertical='center')

        if fill is None:
            fill = PatternFill(fill_type='solid', start_color='0000CCFF', end_color="0000CCFF")

        self.font = font
        self.fill = fill
        self.alignment = alignment

    def format(self, address: Union[str, Tuple[int, int]]) -> "Formatter":
        if isinstance(address, str):
            rng = self.sht[address]
        elif isinstance(address, tuple) and len(tuple) == 2:
            rng = self.sht.cell(address[0], address[1])
        else:
            raise ValueError("address must be a string or tuple")

        rng.font = self.font
        rng.alignment = self.alignment
        rng.fill = self.fill

        return self
